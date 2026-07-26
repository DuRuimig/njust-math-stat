from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


SOURCES = {
    "数学与应用数学": Path("/Users/drm/Documents/Codex/2026-07-18/new-chat/outputs/数学与应用数学专业课程清单.xlsx"),
    "信息与计算科学": Path("/Users/drm/Documents/Codex/2026-07-18/new-chat/outputs/信息与计算科学专业课程清单.xlsx"),
    "应用统计学": Path("/Users/drm/Documents/Codex/2026-07-18/new-chat/outputs/应用统计学专业课程清单.xlsx"),
}
HEADERS = ["semester", "stage", "season", "code", "name", "credits", "category", "requirement", "interdisciplinary", "offering_unit"]


def load_courses(major: str, source: Path) -> list[dict[str, object]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["专业课清单"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    courses = []
    for row in rows:
        if not any(value is not None for value in row):
            continue
        course = dict(zip(HEADERS, row, strict=True))
        course["major"] = major
        course["hours"] = None
        course["teacher"] = None
        courses.append(course)
    return courses


def main() -> None:
    all_courses = [course for major, source in SOURCES.items() for course in load_courses(major, source)]
    summary = {}
    for major in SOURCES:
        courses = [course for course in all_courses if course["major"] == major]
        summary[major] = {
            "course_count": len(courses),
            "credits_in_course_pool": sum(course["credits"] for course in courses),
            "by_category": dict(Counter(course["category"] for course in courses)),
            "by_requirement": dict(Counter(course["requirement"] for course in courses)),
            "by_semester": {str(semester): len(items) for semester, items in sorted(group_by(courses, "semester").items())},
        }
    payload = {
        "source_files": {major: str(source) for major, source in SOURCES.items()},
        "field_notes": {
            "hours": "待补充：源表未提供学时",
            "teacher": "待后续带教师版本提供；本次未推断",
        },
        "summary": summary,
        "courses": all_courses,
    }
    Path("work/course-parse.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_review_markdown(payload)


def group_by(items: list[dict[str, object]], key: str) -> dict[object, list[dict[str, object]]]:
    grouped: dict[object, list[dict[str, object]]] = defaultdict(list)
    for item in items:
        grouped[item[key]].append(item)
    return grouped


def write_review_markdown(payload: dict[str, object]) -> None:
    lines = [
        "# 南理工数统学院专业课解析待确认", "",
        "来源：用户提供的三份专业课程清单 Excel。", "",
        "说明：本文件仅整理课程定义，未包含教师或实际开课安排。学时在原表中未提供，统一标为“待补充”。课程池总学分不等于毕业时实际需修学分。", "",
    ]
    summary = payload["summary"]
    courses = payload["courses"]
    for major, details in summary.items():
        lines.extend([
            f"## {major}", "",
            f"- 课程数：{details['course_count']}；课程池总学分：{details['credits_in_course_pool']}。",
            f"- 模块课程数：{', '.join(f'{category} {count}' for category, count in details['by_category'].items())}。",
            f"- 修读性质：必修 {details['by_requirement'].get('必修', 0)} 门，选修 {details['by_requirement'].get('选修', 0)} 门。", "",
            "| 建议修读学期 | 课程代码 | 课程名称 | 学分 | 专业课分类 | 修读性质 | 交叉融合课 | 学时 |", "| --- | --- | --- | ---: | --- | --- | --- | --- |",
        ])
        major_courses = [course for course in courses if course["major"] == major]
        for course in major_courses:
            lines.append(
                f"| {course['semester']}（{course['stage']}，{course['season']}） | {course['code']} | {course['name']} | {course['credits']} | {course['category']} | {course['requirement']} | {course['interdisciplinary']} | 待补充 |"
            )
        lines.append("")
    Path("work/专业课解析待确认.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
